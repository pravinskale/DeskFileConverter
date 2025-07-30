from datetime import datetime
import re
import pdfplumber as pdfplumber
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import json
    
#pdfToConvert = "C:\path\Docs\Test_Monthly_Report_All_2025-07-08_14_49_40.pdf"
#convertedExcel = "C:\path\Docs\Test_Monthly_Report_All_2025-07-08_14_49_40.xlsx"

class PDFTOExcelConverter:
    column_count=0;
    excel_headers = []
    gto_takeoffs = []
    resolution_areas ={}
    assignment_group = []
    def __init__(self):
        """
        Initialize the PDFTOExcelConverter class.
        """
        #Read report_config.json file and get the GTO values
        try:
            with open('\\\\USLTCSBMPSVC01\\Share\\Pravin\\PDFToExl\\report_config.json', 'r') as file:
                config = json.load(file)
                self.gto_takeoffs = config.get("GTO", [])
                self.excel_headers = config.get("ExcelHeaders", [])
                self.resolution_areas = config.get("ResolutionAreas", {})
                self.assignment_groups = config.get("AssignmentGroups", [])

        except FileNotFoundError:
            print("report_config.json file not found. Using default GTO values.")
        except json.JSONDecodeError:
            print("Error decoding JSON from report_config.json. Using default GTO values.")
    
    def get_nxt_col_idx(self):
        PDFTOExcelConverter.column_count = PDFTOExcelConverter.column_count + 1;
        return PDFTOExcelConverter.column_count;
    #open the excel file and auto adjust the column width
    def format_header_row(self, ws):
        """ Adjust the width of the columns in the Excel file based on the content and also format the header row.
            This function iterates through each column in the worksheet and sets the width based on the maximum
        """
        header_front = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        for column in ws.columns:
            max_length = 0
            fixed_length = 40
            column_letter = get_column_letter(column[0].column)  # Get the column letter
            for cell in column:
                try:
                    cell_len = len(str(cell.value))
                    #max_length = max(max_length, len(str(cell.value)))
                    if cell_len > fixed_length:
                        max_length = fixed_length
                    else:
                        max_length = max(max_length, len(str(cell.value)))
                    # Apply header formatting
                    if cell.row == 1:  # Header row
                        cell.font = header_front
                        cell.fill = header_fill
                        cell.border = header_border

                except Exception as e:
                    print(f"Error processing cell '{cell.coordinate}': {e}")
            adjusted_width = (max_length + 2)  # Add some padding
            ws.column_dimensions[column_letter].width = adjusted_width
            

    # Open the converted Excel file and update the column header names
    def update_column_headers(self, ws):
        # Update the first row with new headers
        
        PDFTOExcelConverter.column_count = len(self.excel_headers);

        for col_num, header in enumerate(self.excel_headers, start=1):
            ws.cell(row=1, column=col_num, value=header)
    
    def add_res_area_column(self, workSheet):
        #find last column index in the worksheet
        res_area_idx = self.get_nxt_col_idx();
        workSheet.insert_cols(res_area_idx)  # Insert a new column at position 21 (after "Tags")
        workSheet.cell(row=1, column=res_area_idx, value="Resolution Area")

        tkff_area_idx = self.get_nxt_col_idx();
        workSheet.insert_cols(tkff_area_idx)  # Insert a new column at position 21 (after "Tags")
        workSheet.cell(row=1, column=tkff_area_idx, value="Takeoff")

        # Populate the new column with values from the res_area dictionary
        for row in range(2, workSheet.max_row + 1):
            tags_value = workSheet.cell(row=row, column=ExcelHeaders.Tags.value).value  # Assuming "Tags" is in column 16
            #Split the tags value by comma and if there are multiple values, then add second coloumn
            if(tags_value):
                tags_value = tags_value.split(',')
                if len(tags_value) > 0:
                    for i, tag in enumerate(tags_value):
                        tag = tag.strip()
                        if(tag.startswith("CO_L3") or tag.startswith("TO_L3")):
                            if tag in self.resolution_areas:
                                workSheet.cell(row=row, column=res_area_idx, value=self.resolution_areas[tag])  # Set the value in the new column
                            else:
                                workSheet.cell(row=row, column=res_area_idx, value=tag)  # Default value if not found
                        else:
                            workSheet.cell(row=row, column=tkff_area_idx, value=tag)
        self.add_GTO_column(workSheet, tkff_area_idx);
    
    # Add GTO column after the resolution area column
    # This function adds a new column for GTO values in the Excel file.    
    def add_GTO_column(self, workSheet, tkff_area_idx):
        #add a new column in the excel for GTO column
        gto_idx = self.get_nxt_col_idx();
        workSheet.insert_cols(gto_idx)  # Insert a new column at position 22
        workSheet.cell(row=1, column=gto_idx, value="GTO")
        for row in range(2, workSheet.max_row + 1):
            tkff_value = workSheet.cell(row=row, column=tkff_area_idx).value
            tkff_value = tkff_value.replace(" ", "") if tkff_value else ""  
            if tkff_value and tkff_value.strip() in self.gto_takeoffs:
                workSheet.cell(row=row, column=gto_idx, value="GTO")
            else:
                workSheet.cell(row=row, column=gto_idx, value="#N/A")

    def add_month_column(self, workSheet):
        """
        Create a new column in the Excel file to indicate the month of the created date.
        """
        month_idx = self.get_nxt_col_idx();
        workSheet.insert_cols(month_idx)  # Insert a new column at position 24
        workSheet.cell(row=1, column=month_idx, value="Month")
        # Populate the new column with the month from the "Created" date
        for row in range(2, workSheet.max_row + 1):
            created_date = workSheet.cell(row=row, column=ExcelHeaders.CreatedDate.value).value  # Assuming "Created" is in column 13
            if created_date:
                try:
                    # Convert the created date to a datetime object
                    created_date = pd.to_datetime(self.format_date(created_date), errors='coerce')
                    if pd.notna(created_date):
                        month_name = created_date.strftime("%B")  # Get the full month name
                        workSheet.cell(row=row, column=month_idx, value=month_name)
                    else:
                        workSheet.cell(row=row, column=month_idx, value="")
                except Exception as e:
                    print(f"Month Column: Error processing date '{created_date}': {e}")
                    workSheet.cell(row=row, column=month_idx, value="NA")

    def add_res_days_column(self, workSheet):
        """
        Add a new column in the Excel file to calculate the number of days from the created date to the end of the month.
        """
        res_days_idx = self.get_nxt_col_idx();
        workSheet.insert_cols(res_days_idx)
        workSheet.cell(row=1, column=res_days_idx, value="Res 15")
        # Populate the new column with the number of days subtracted from Resolved date and created date
        for row in range(2, workSheet.max_row + 1):
            created_date = workSheet.cell(row=row, column=ExcelHeaders.CreatedDate.value).value
            resolved_date = workSheet.cell(row=row, column=ExcelHeaders.ResolvedDate.value).value  # Assuming "Resolved" is in column 13
            if created_date and resolved_date:
                try:
                    # Convert the created date to a datetime object
                    created_date = pd.to_datetime(self.format_date(created_date), errors='coerce')
                    resolved_date = pd.to_datetime(self.format_date(resolved_date), errors='coerce')
                    if pd.notna(created_date) and pd.notna(resolved_date):
                        res_days = (resolved_date - created_date).days
                        workSheet.cell(row=row, column=res_days_idx, value=res_days)
                    else:
                        workSheet.cell(row=row, column=res_days_idx, value="")
                except Exception as e:
                    print(f"Res Days: Error processing dates '{created_date}' and '{resolved_date}': {e}")
                    workSheet.cell(row=row, column=res_days_idx, value="")

    def add_days_post_submission_column(self, workSheet):
        """
        Add a new column in the Excel file to calculate the number of days from the created date to the current date.
        """
        days_post_submission_idx = self.get_nxt_col_idx();
        workSheet.insert_cols(days_post_submission_idx)
        workSheet.cell(row=1, column=days_post_submission_idx, value="Days Post Submission")
        # Populate the new column with the number of days from created date to current date
        for row in range(2, workSheet.max_row + 1):
            created_date = workSheet.cell(row=row, column=ExcelHeaders.CreatedDate.value).value #Created date column is at index 10
            if created_date:
                try:
                    # Convert the created date to a datetime object
                    created_date = pd.to_datetime(self.format_date(created_date), errors='coerce')
                    if pd.notna(created_date):
                        current_date = pd.to_datetime(datetime.now().strftime("%Y-%m-%d"), errors='coerce')
                        days_post_submission = (current_date - created_date).days
                        workSheet.cell(row=row, column=days_post_submission_idx, value=days_post_submission)
                    else:
                        workSheet.cell(row=row, column=days_post_submission_idx, value="")
                except Exception as e:
                    print(f"Post Submission: Error processing date '{created_date}': {e}")
                    workSheet.cell(row=row, column=days_post_submission_idx, value="")

    def add_description_column(self, workSheet):
        """
        Add a new column in the Excel file to provide a description based on the "Short description" column.
        """
        description_idx = self.get_nxt_col_idx();
        workSheet.insert_cols(description_idx)
        workSheet.cell(row=1, column=description_idx, value="Description")
        # Populate the new column with a description based on the "Short description" column
        for row in range(2, workSheet.max_row + 1):
            short_description = workSheet.cell(row=row, column=ExcelHeaders.ShortDesc.value).value
            if short_description:
                try:
                    # Create a description based on the short description
                    short_description = re.sub(r'[\n\r\t\f]+', ' ', short_description)  # Clean up the short description
                    #Find first occurence of a hyphen and split the string
                    description ="";
                    if '-' in short_description:
                        description = short_description.split('-')[0].strip()

                    workSheet.cell(row=row, column=description_idx, value=description)
                except Exception as e:
                    print(f"Error processing short description '{short_description}': {e}")
                    workSheet.cell(row=row, column=description_idx, value="Error")

    def add_agging_column(self, workSheet):
        """
        Add a new column in the Excel file to calculate the aging of the task based on the created date.
        """
        aging_idx = self.get_nxt_col_idx();
        workSheet.insert_cols(aging_idx)
        workSheet.cell(row=1, column=aging_idx, value="Aging")
        # Populate the new column with the aging information
        for row in range(2, workSheet.max_row + 1):
            created_date = workSheet.cell(row=row, column=ExcelHeaders.CreatedDate.value).value
            try:
                created_date = self.format_date(created_date)
                if pd.notna(created_date):
                    # Calculate aging based on the end of previous month
                    end_of_month = (pd.Timestamp.today() + pd.offsets.MonthEnd(-1)).normalize()
                    aging = (end_of_month - created_date).days
                else:
                    aging = ""
                workSheet.cell(row=row, column=aging_idx, value=aging)
            except Exception as e:
                print(f"PDF To Data Grid: Error processing date '{created_date}': {e}")
    def format_date(self, date_str, isPrint=False):
        """
        Format the date in the Excel file to a specific format.
        """
        if not date_str:
            raise ValueError("Date string is empty or None")
        date_str = date_str.replace("- ", "-")
        date_str = re.match(r"\d{2}-[A-Za-z]{3}-\d{2}", date_str).group()
        if isPrint: print(f"Created Date before->" + date_str)
        #Convert to datetime object
        date_obj = datetime.strptime(date_str, "%d-%b-%y")
        formatted_date = pd.to_datetime(date_obj.strftime("%Y-%m-%d"), errors='coerce')
        if isPrint: print(f"Created Date after->{formatted_date}")
        return formatted_date if pd.notna(formatted_date) else None
    
    def format_name_description(self, description):
        """
        Format the description in the Excel file to a specific format.
        """
        if not description:
            return ""
        description = description.strip()
        # Replace spaces not followed by uppercase letters
        formatted = re.sub(r' (?=[^A-Z])', '', description)
        return formatted
    
    def format_assignment_group(self, assignment_group):
        """
        Format assignment group as per the AssignmentGroups values in the config file
        """
        if not assignment_group:
            return ""
        assignment_group = assignment_group.replace(" ", "")
        if assignment_group in self.assignment_groups:
            return self.assignment_groups[assignment_group]
        return "<Missing Group>"

    def pdf_to_excel(self, pdfToConvert, convertedExcel):
        """
        Convert PDF to Excel and add aging information based on the created date.
        """
        all_data = []
         # Iterate through the rows and add aging information
        created_date_idx = ExcelHeaders.CreatedDate.value - 1 #For PDF column index starts at 0
        customer_name_idx = ExcelHeaders.Customer.value - 1
        assigned_to_idx = ExcelHeaders.AssignedTo.value - 1
        assigned_grp_idx = ExcelHeaders.AssignmentGroup.value - 1
        catalog_task_idx = ExcelHeaders.CatalogTask.value - 1
        task_idx = ExcelHeaders.Task.value - 1
        parent_idx = ExcelHeaders.Parent.value - 1
        tags_idx = ExcelHeaders.Tags.value - 1
        task_type_idx = ExcelHeaders.TaskType.value - 1
        state_idx = ExcelHeaders.State.value - 1

        with pdfplumber.open(pdfToConvert) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if table:
                    cleaned_table = []
                    for row in table:
                        cleaned_row = [re.sub(r'[\n\r\t\f]+', ' ', cell) if cell else '' for cell in row]
                        if(row[0].startswith("Run By :") == False):
                            cleaned_table.append(cleaned_row)
                    # Extract header and ensure uniqueness
                    header = cleaned_table[0]
                    seen = {}
                    unique_header = []
                    for col in header:
                        #Ignore empty columns
                        # if col.strip() == "":
                        #     print(f"--{col}")
                        #     continue;         
                        if col in seen:
                            seen[col] += 1
                            unique_header.append(f"{col}_{seen[col]}")
                        else:
                            seen[col] = 0
                            unique_header.append(col)

                    for row in cleaned_table[1:]:
                        #format customer name
                        if len(row) > customer_name_idx and row[customer_name_idx]:
                            row[customer_name_idx] = self.format_name_description(row[customer_name_idx])
                        
                        #Format assigned to column
                        if len(row) > assigned_to_idx and row[assigned_to_idx]:
                            row[assigned_to_idx] = self.format_name_description(row[assigned_to_idx]);
                        #Format Assigned group column
                        if  len(row) >assigned_grp_idx and row[assigned_grp_idx]:
                            #format assigment group
                            row[assigned_grp_idx] = self.format_assignment_group(row[assigned_grp_idx])#row[assigned_grp_idx].replace(" ", "")

                        if  len(row) >catalog_task_idx and row[catalog_task_idx]:
                            row[catalog_task_idx] = row[catalog_task_idx].replace(" ", "")
                        
                        if  len(row) >task_idx and row[task_idx]:
                            row[task_idx] = row[task_idx].replace(" ", "")
                        if  len(row) >parent_idx and row[parent_idx]:
                            row[parent_idx] = row[parent_idx].replace(" ", "")

                        #Format created date
                        if  len(row) >created_date_idx and row[created_date_idx]:
                            row[created_date_idx] = row[created_date_idx].replace("- ", "-").replace(" -", "-")
                        #Format tags column replace space
                        if len(row) > tags_idx and row[tags_idx]:
                            #print(row[tags_idx]);
                            row[tags_idx] = row[tags_idx].replace(" ", "")
                        
                        if len(row) > task_type_idx and row[task_type_idx]:
                            # Format task type to remove spaces
                            row[task_type_idx] = self.format_name_description(row[task_type_idx])
                        if len(row) > state_idx and row[state_idx]: 
                            # Format state to remove spaces
                            row[state_idx] = self.format_name_description(row[state_idx])

                        # Remove blank last columns if they exist to make sure it matches the headers
                        #header_row_count_diff = len(row) - len(unique_header)
                        #if header_row_count_diff > 0:
                         #   del row[-(header_row_count_diff):]

                    #print(f"Header column {len(unique_header)} and non header first row {len(cleaned_table[1])}")
                    df = pd.DataFrame(cleaned_table[1:], columns=unique_header)  # Skip header row
                    all_data.append(df)

        # Combine all pages into one DataFrame
        final_df = pd.concat(all_data, ignore_index=True)

        # Save to Excel
        final_df.to_excel(convertedExcel, index=False)
        wb = openpyxl.load_workbook(convertedExcel)
        workSheet = wb.active

        self.update_column_headers(workSheet)
        self.add_agging_column(workSheet)
        self.add_res_area_column(workSheet)
        self.add_month_column(workSheet)
        self.add_res_days_column(workSheet)
        self.add_days_post_submission_column(workSheet);
        self.add_description_column(workSheet);
        
        self.format_header_row(workSheet)
        wb.save(convertedExcel)
        #Add a function create a new column in the excel file
        
from enum import Enum
class ExcelHeaders(Enum):
    TaskType = 1
    CatalogTask = 2
    Parent = 3
    Task = 4
    Customer = 5
    ShortDesc = 6
    State = 7
    AssignmentGroup = 8
    AssignedTo = 9
    CreatedDate = 10
    Priority = 11
    ClosedDate = 12
    ResolvedDate = 13
    UpdatedDate = 14
    HasBreached = 15
    Tags = 16
    ChildIncident = 17
    ParentIncident = 18
